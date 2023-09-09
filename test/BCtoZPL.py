from turtle import width
import openpyxl
import os
import zpl

def main():

    for file in os.listdir(os.fsencode("./")):
        filename = os.fsdecode(file)

        if filename.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(filename)
            for worksheet in workbook.worksheets:

                for row in range(1, worksheet.max_row):
                    for column in worksheet.iter_cols(1, 1):
                        value = column[row].value
                        
                        zbc = zpl.Label(50.8, 152.4)
                        zbc.origin(20, 20)
                        zbc.barcode("C", value, height=300, width=300)
                        zbc.endorigin()
                        
                        print(zbc.dumpZPL())
                        zbc.preview()


if __name__ == '__main__':
    main()