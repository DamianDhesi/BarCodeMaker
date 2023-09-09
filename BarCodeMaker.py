from barcode import Code128
from barcode.writer import ImageWriter
import openpyxl
import os

def main():

    if not os.path.exists("./BarCodes"):
      os.mkdir("./BarCodes")

    barcodewriter = ImageWriter()
    barcodewriter.font_path = "./DejaVuSansMono.ttf"

    for file in os.listdir(os.fsencode("./")):
        filename = os.fsdecode(file)
        
        if filename.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(filename)
            for worksheet in workbook.worksheets:
    
                for row in range(1, worksheet.max_row):
                    for column in worksheet.iter_cols(1, 1):
                        value1 = column[row].value

                    barcode1 = Code128(value1, writer=barcodewriter)

                    options = dict(module_width=0.578, 
                                   module_height=23.9, font_size=17)

                    barcode1.save(os.path.join("./BarCodes/", value1), options)

if __name__ == '__main__':
    main()